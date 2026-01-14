# excel_exporter.py

import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ======================================================
# 🔧 EXPORTADOR EXCEL FORMATEADO
# ======================================================

def exportar_excel_formateado(df, archivo, nombre_reporte="Reporte"):
    """Crea un Excel con una hoja por tienda, con formato visual profesional."""

    if df.empty:
        raise ValueError("El DataFrame está vacío, no se puede exportar.")

    # 🔍 Detectar columna de tienda según el tipo de reporte
    if "tienda" in df.columns:
        col_tienda = "tienda"
    elif "tienda_origen" in df.columns:
        col_tienda = "tienda_origen"
    elif "tienda_destino" in df.columns:
        col_tienda = "tienda_destino"
    else:
        raise KeyError(
            f"No se encontró ninguna columna de tienda en el DataFrame.\n"
            f"Columnas disponibles: {list(df.columns)}"
        )

    # --- Crear archivo base ---
    with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
        for tienda, df_tienda in df.groupby(col_tienda, sort=True):
            hoja = tienda[:25] if isinstance(tienda, str) else str(tienda)
            df_tienda.to_excel(writer, sheet_name=hoja, index=False)

    wb = load_workbook(archivo)
    thin = Side(border_style="thin", color="000000")

    # 🎨 Colores de estilo
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    alt_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")

    for ws in wb.worksheets:
        ws.row_dimensions[1].height = 40
        for cell in ws[1]:
            cell.font = Font(bold=True, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.fill = header_fill

        # --- Ajuste de anchos ---
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in ["E", "F", "G", "H"]:
                if col_letter == "E":
                    ws.column_dimensions[col_letter].width = 56 / 7
                elif col_letter == "F":
                    ws.column_dimensions[col_letter].width = 46 / 7
                elif col_letter == "G":
                    ws.column_dimensions[col_letter].width = 94 / 7
                elif col_letter == "H":
                    ws.column_dimensions[col_letter].width = 81 / 7
            else:
                max_length = 0
                for cell in ws[get_column_letter(col_idx)]:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = (max_length + 2) * 1.1

        # --- Bordes, alineación y alternancia de color ---
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), start=2):
            fill = alt_fill if i % 2 == 0 else None
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                if fill:
                    cell.fill = fill

        ws.freeze_panes = "A2"

        # --- Configuración de impresión ---
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5)

        # --- Encabezado / pie ---
        try:
            ws.HeaderFooter.left_header.text = f"&LJAGI - {nombre_reporte}"
            ws.HeaderFooter.right_header.text = "&R&P de &N"
            ws.HeaderFooter.center_footer.text = "&CGenerado automáticamente"
        except AttributeError:
            pass

    wb.save(archivo)
    print(f"\n🖨️ Archivo listo y formateado: {archivo}")